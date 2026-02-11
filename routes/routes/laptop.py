import io
import base64
from datetime import datetime, date

from io import BytesIO

from flask import Blueprint, render_template, request, make_response
from flask_login import login_required
from sqlalchemy import or_

from extensions import db
from models import LaptopLog

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from reportlab.lib.units import mm

from fingerprint.factory import get_fingerprint_service

laptop_bp = Blueprint("laptop", __name__, url_prefix="/laptops")

# Daftar nama laptop (nanti boleh ganti sesuai aset beneran)
LAPTOP_NAMES = [
    "LAPTOP-NOC-01",
]

# Daftar nama petugas (reuse dari kunci)
STAFF_NAMES = [
    "Dhanu",
    "Ariq",
    "Rendy",
    "Mulya",
    "Alex",
    "Frengki",
    "Ardan",
    "Rio",
    "Jovi",
    "Ali",
]

def get_return_status(log):
    """Hitung apakah pengembalian tepat waktu / telat / belum kembali."""
    if not log.planned_return_date:
        return "-"

    planned = log.planned_return_date
    today = date.today()

    # Kalau SUDAH kembali
    if log.time_in:
        actual = log.time_in.date()
        if actual <= planned:
            return "Tepat waktu"
        else:
            diff = (actual - planned).days
            return f"Terlambat {diff} hari"

    # Kalau BELUM kembali
    if today < planned:
        diff = (planned - today).days
        return f"Belum kembali (masih {diff} hari)"
    elif today == planned:
        return "Belum kembali (jatuh tempo hari ini)"
    else:
        diff = (today - planned).days
        return f"Belum kembali (terlambat {diff} hari)"


def get_filtered_laptops(args):
    """Helper untuk filter/search log laptop (mirip key_logs)."""
    q = LaptopLog.query

    text = (args.get("q") or "").strip()
    status = (args.get("status") or "").strip()
    date_from_str = (args.get("date_from") or "").strip()
    date_to_str = (args.get("date_to") or "").strip()

    # Search text: nama, divisi, laptop, keperluan, ticket
    if text:
        like = f"%{text}%"
        q = q.filter(
            or_(
                LaptopLog.borrower_name.ilike(like),
                LaptopLog.borrower_division.ilike(like),
                LaptopLog.laptop_name.ilike(like),
                LaptopLog.purpose.ilike(like),
                LaptopLog.ticket_mbs.ilike(like),
            )
        )

    # Filter status
    if status in ("borrowed", "returned"):
        q = q.filter_by(status=status)

    # Filter tanggal pakai Tanggal Pinjam (time_out)
    if date_from_str:
        try:
            df = datetime.strptime(date_from_str, "%Y-%m-%d").date()
            start_dt = datetime.combine(df, datetime.min.time())
            q = q.filter(LaptopLog.time_out >= start_dt)
        except ValueError:
            pass

    if date_to_str:
        try:
            dt_ = datetime.strptime(date_to_str, "%Y-%m-%d").date()
            end_dt = datetime.combine(dt_, datetime.max.time())
            q = q.filter(LaptopLog.time_out <= end_dt)
        except ValueError:
            pass

    # urutkan dari pinjam terbaru
    q = q.order_by(LaptopLog.time_out.desc().nullslast())
    return q.all()



@laptop_bp.route("/")
@login_required
def index():
    logs = get_filtered_laptops(request.args)

    # laptop yang sedang dipinjam (status = borrowed)
    borrowed_logs = (
        LaptopLog.query
        .filter_by(status="borrowed")
        .order_by(LaptopLog.laptop_name.asc())
        .all()
    )

    return render_template(
        "laptop/index.html",
        logs=logs,
        borrowed_logs=borrowed_logs,
        get_return_status=get_return_status,
    )



@laptop_bp.route("/data")
@login_required
def data():
    """Dipakai HTMX buat refresh tabel laptop."""
    logs = get_filtered_laptops(request.args)
    return render_template(
        "laptop/_table_body.html",
        logs=logs,
        get_return_status=get_return_status,
    )



@laptop_bp.route("/borrow-form")
@login_required
def borrow_form():
    """Tampilkan modal peminjaman laptop."""
    now = datetime.now()

    # cari laptop yang lagi borrowed
    borrowed_rows = (
        db.session.query(LaptopLog.laptop_name)
        .filter(
            LaptopLog.status == "borrowed",
            LaptopLog.laptop_name.in_(LAPTOP_NAMES),
        )
        .all()
    )
    borrowed_set = {row[0] for row in borrowed_rows}

    laptops = [
        {"name": name, "is_borrowed": name in borrowed_set}
        for name in LAPTOP_NAMES
    ]

    return render_template(
        "laptop/_borrow_modal.html",
        now=now,
        laptops=laptops,
        staff_names=STAFF_NAMES,
    )


@laptop_bp.route("/borrow", methods=["POST"])
@login_required
def borrow():
    
    """Proses form peminjaman laptop."""
    has_ticket = request.form.get("has_ticket")  # "yes" atau "no"
    ticket_mbs = (request.form.get("ticket_mbs") or "").strip()

    borrower_name = (request.form.get("borrower_name") or "").strip()
    borrower_division = (request.form.get("borrower_division") or "").strip()
    purpose = (request.form.get("purpose") or "").strip()
    laptop_name = request.form.get("laptop_name")
    time_out_str = request.form.get("time_out")
    planned_return_str = (request.form.get("planned_return_date") or "").strip()
    staff_out_name = (request.form.get("staff_out_name") or "").strip()

    sig_in_visitor = request.form.get("sig_in_visitor")
    sig_in_staff = request.form.get("sig_in_staff")

    # ===== Validasi basic =====
    if not borrower_name or not borrower_division or not purpose or not laptop_name or not time_out_str or not staff_out_name:
        logs = get_filtered_laptops(request.args)
        return render_template(
            "laptop/_table_body.html",
            logs=logs,
            get_return_status=get_return_status,
        ), 400
    # Kalau pilih YES ticket, tapi kosong → tolak
    if has_ticket == "yes" and not ticket_mbs:
        logs = get_filtered_laptops(request.args)
        return render_template(
            "laptop/_table_body.html",
            logs=logs,
            get_return_status=get_return_status,
        ), 400
    # TTD wajib
    if not sig_in_visitor or not sig_in_staff:
        logs = get_filtered_laptops(request.args)
        return render_template(
            "laptop/_table_body.html",
            logs=logs,
            get_return_status=get_return_status,
        ), 400
    # Parse tanggal pinjam
    try:
        time_out = datetime.strptime(time_out_str, "%Y-%m-%dT%H:%M")
    except ValueError:
        time_out = datetime.now()

    # Parse rencana tanggal kembali (boleh kosong)
    planned_return_date = None
    if planned_return_str:
        try:
            planned_return_date = datetime.strptime(planned_return_str, "%Y-%m-%d").date()
        except ValueError:
            planned_return_date = None

    # Cek: laptop ini sudah dipinjam orang lain belum?
    existing = LaptopLog.query.filter_by(
        laptop_name=laptop_name,
        status="borrowed",
    ).first()
    if existing:
        logs = get_filtered_laptops(request.args)
        return render_template(
            "laptop/_table_body.html",
            logs=logs,
            get_return_status=get_return_status,
        ), 400
    
        # ===== Fingerprint verification (PINJAM) =====
    fp = get_fingerprint_service()
    result = fp.verify(staff_out_name)

    if not result.get("matched"):
        logs = get_filtered_laptops(request.args)
        return render_template(
            "laptop/_table_body.html",
            logs=logs,
            get_return_status=get_return_status,
        ), 403


    # ===== Simpan ke DB =====
    log = LaptopLog(
        ticket_mbs=ticket_mbs if has_ticket == "yes" else None,
        borrower_name=borrower_name,
        borrower_division=borrower_division,
        purpose=purpose,
        laptop_name=laptop_name,
        time_out=time_out,
        planned_return_date=planned_return_date,
        status="borrowed",
        signature_out_borrower=sig_in_visitor,
        signature_out_staff=sig_in_staff,
        staff_out_name=staff_out_name,
    )

    db.session.add(log)
    db.session.commit()

    logs = get_filtered_laptops(request.args)
    return render_template(
        "laptop/_table_body.html",
        logs=logs,
        get_return_status=get_return_status,
    )

@laptop_bp.route("/detail/<int:log_id>")
@login_required
def detail(log_id):
    log = LaptopLog.query.get_or_404(log_id)
    return render_template(
        "laptop/_detail_modal.html",
        log=log,
        get_return_status=get_return_status,
    )

@laptop_bp.route("/return-form/<int:log_id>")
@login_required
def return_form(log_id):
    log = LaptopLog.query.get_or_404(log_id)
    now = datetime.now()
    return render_template(
        "laptop/_return_modal.html",
        log=log,
        now=now,
        staff_names=STAFF_NAMES,
    )

@laptop_bp.route("/return/<int:log_id>", methods=["POST"])
@login_required
def return_laptop(log_id):
    log = LaptopLog.query.get_or_404(log_id)

    time_in_str = request.form.get("time_in")
    staff_in_name = (request.form.get("staff_in_name") or "").strip()
    sig_out_visitor = request.form.get("sig_out_visitor")
    sig_out_staff = request.form.get("sig_out_staff")

    # validasi basic
    if not time_in_str or not staff_in_name:
        logs = get_filtered_laptops(request.args)
        return render_template(
            "laptop/_table_body.html",
            logs=logs,
            get_return_status=get_return_status,
        ), 400
    if not sig_out_visitor or not sig_out_staff:
        logs = get_filtered_laptops(request.args)
        return render_template(
            "laptop/_table_body.html",
            logs=logs,
            get_return_status=get_return_status,
        ), 400
    try:
        time_in = datetime.strptime(time_in_str, "%Y-%m-%dT%H:%M")
    except ValueError:
        time_in = datetime.now()

        # ===== Fingerprint verification (KEMBALI) =====
    fp = get_fingerprint_service()
    result = fp.verify(staff_in_name)

    if not result.get("matched"):
        logs = get_filtered_laptops(request.args)
        return render_template(
            "laptop/_table_body.html",
            logs=logs,
            get_return_status=get_return_status,
        ), 403


    log.time_in = time_in
    log.status = "returned"
    log.signature_in_borrower = sig_out_visitor
    log.signature_in_staff = sig_out_staff
    log.staff_in_name = staff_in_name

    db.session.commit()

    logs = get_filtered_laptops(request.args)
    return render_template(
        "laptop/_table_body.html",
        logs=logs,
        get_return_status=get_return_status,
    )
@laptop_bp.route("/export/excel")
@login_required
def export_excel():
    """Export log laptop ke Excel sesuai filter di halaman."""
    logs = get_filtered_laptops(request.args)

    wb = Workbook()
    ws = wb.active
    ws.title = "Log Laptop"

    # Header
    headers = [
        "ID",
        "Ticket MBS",
        "Peminjam",
        "Divisi",
        "Keperluan",
        "Laptop",
        "Status",
        "Tanggal Pinjam",
        "Tanggal Kembali",
        "Rencana Tgl Kembali",
        "Petugas PINJAM",
        "Petugas KEMBALI",
        "Dibuat",
        "Diupdate",
    ]

    ws.append(headers)

    # Styling header
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.alignment = header_alignment

    # Data rows
    for log in logs:
        row = [
            log.id,
            log.ticket_mbs or "",
            log.borrower_name or "",
            log.borrower_division or "",
            log.purpose or "",
            log.laptop_name or "",
            log.status or "",
            log.time_out.strftime("%Y-%m-%d %H:%M") if log.time_out else "",
            log.time_in.strftime("%Y-%m-%d %H:%M") if log.time_in else "",
            log.planned_return_date.strftime("%Y-%m-%d") if log.planned_return_date else "",
            log.staff_out_name or "",
            log.staff_in_name or "",
            log.created_at.strftime("%Y-%m-%d %H:%M") if log.created_at else "",
            log.updated_at.strftime("%Y-%m-%d %H:%M") if log.updated_at else "",
        ]
        ws.append(row)

    # Auto width kolom
    for col_idx, _ in enumerate(headers, start=1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in ws[column_letter]:
            try:
                val = str(cell.value) if cell.value is not None else ""
            except Exception:
                val = ""
            max_length = max(max_length, len(val))
        ws.column_dimensions[column_letter].width = max_length + 2

    # Simpan ke memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"laptop-log-{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"

    response = make_response(output.getvalue())
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    response.headers["Content-Type"] = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    return response

@laptop_bp.route("/export/pdf")
@login_required
def export_pdf():
    logs = get_filtered_laptops(request.args)

    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    def header():
        c.setFont("Helvetica-Bold", 14)
        c.drawString(20*mm, height - 20*mm, "Laporan Peminjaman Laptop")
        c.setFont("Helvetica", 9)
        c.drawString(20*mm, height - 26*mm, f"Dibuat: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        c.line(20*mm, height - 28*mm, width - 20*mm, height - 28*mm)

    def sig_box(label, sig, x, y_box, box_w, box_h):
        c.setFont("Helvetica", 8)
        c.drawString(x, y_box + box_h + 2*mm, label)
        c.rect(x, y_box, box_w, box_h)
        if not sig:
            c.setFont("Helvetica-Oblique", 7)
            c.drawString(x+2*mm, y_box+box_h/2, "Tidak ada TTD.")
            return
        
        try:
            sig_data = sig.split(",", 1)[1]
            img = ImageReader(io.BytesIO(base64.b64decode(sig_data)))
            c.drawImage(img, x+2*mm, y_box+2*mm, width=box_w-4*mm, height=box_h-4*mm,
                        preserveAspectRatio=True, anchor="c")
        except:
            c.setFont("Helvetica-Oblique", 7)
            c.drawString(x+2*mm, y_box+box_h/2, "TTD error.")

    header()
    y = height - 40*mm
    bottom = 30*mm

    if not logs:
        c.setFont("Helvetica-Oblique", 10)
        c.drawString(20*mm, y, "Tidak ada transaksi sesuai filter.")
    else:
        for idx, log in enumerate(logs, start=1):

            if y < bottom:
                c.showPage()
                header()
                y = height - 40*mm

            c.setFont("Helvetica-Bold", 10)
            c.drawString(20*mm, y, f"{idx}. Laptop: {log.laptop_name}")
            y -= 5*mm

            c.setFont("Helvetica", 9)
            c.drawString(20*mm, y, f"Peminjam: {log.borrower_name}")
            y -= 4*mm

            c.drawString(20*mm, y, f"Divisi: {log.borrower_division}")
            y -= 4*mm

            c.drawString(20*mm, y, f"Keperluan: {log.purpose}")
            y -= 4*mm

            c.drawString(20*mm, y, f"Status: {log.status}")
            y -= 4*mm

            c.drawString(20*mm, y, f"Tgl Pinjam: {log.time_out.strftime('%Y-%m-%d %H:%M') if log.time_out else '-'}")
            y -= 4*mm

            c.drawString(20*mm, y, f"Tgl Kembali: {log.time_in.strftime('%Y-%m-%d %H:%M') if log.time_in else '-'}")
            y -= 6*mm

            box_w = (width - 40*mm) / 2
            box_h = 25*mm

            y1 = y - box_h
            if y1 < bottom:
                c.showPage()
                header()
                y = height - 40*mm
                y1 = y - box_h

            sig_box("TTD Peminjam (PINJAM)", log.signature_out_borrower, 20*mm, y1, box_w, box_h)
            sig_box("TTD Petugas (PINJAM)", log.signature_out_staff, 20*mm + box_w, y1, box_w, box_h)

            y = y1 - 7*mm

            y2 = y - box_h
            if y2 < bottom:
                c.showPage()
                header()
                y = height - 40*mm
                y2 = y - box_h

            sig_box("TTD Peminjam (KEMBALI)", log.signature_in_borrower, 20*mm, y2, box_w, box_h)
            sig_box("TTD Petugas (KEMBALI)", log.signature_in_staff, 20*mm + box_w, y2, box_w, box_h)

            y = y2 - 10*mm

    c.save()
    pdf = buffer.getvalue()
    buffer.close()

    filename = f"LaptopLog_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    response = make_response(pdf)
    response.headers["Content-Type"] = "application/pdf"
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response
