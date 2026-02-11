import io
import base64

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from reportlab.lib.units import mm

from datetime import datetime, date, timedelta
from sqlalchemy import or_

from flask import Blueprint, render_template, request, redirect, url_for, abort, make_response
from flask_login import login_required, current_user
from extensions import db
from models import KeyLog

from openpyxl import Workbook 
from openpyxl.styles import Font, Alignment
from fingerprint.factory import get_fingerprint_service


RACK_NAMES = [
    "SUMBING",
    "CIREMAI",
    "ELBRUS",
    "FUJI",
    "LAWU",
    "SEMERU",
    "PAPANDAYAN",
    "MERBABU",
    "RINJANI",
    "ALPEN",
    "MERAPI",
    "BROMO",
    "KILIMANJARO",
    "EVEREST",
]

STAFF_NAMES = [
    "Alex",
    "Ali",
    "Ardan",
    "Ariq",
    "Dhanu",
    "Frengki",
    "Jovi",
    "Mulya",
    "Rendy",
    "Rio",
    
    
]



key_bp = Blueprint("key", __name__, url_prefix="/keys")

def get_filtered_logs(args):
    q = KeyLog.query

    text = (args.get("q") or "").strip()
    status = (args.get("status") or "").strip()
    date_from_str = (args.get("date_from") or "").strip()
    date_to_str = (args.get("date_to") or "").strip()
    sort = (args.get("sort") or "time_in_desc").strip()

    # Search text: ticket, pengunjung, perusahaan, rak
    if text:
        like = f"%{text}%"
        q = q.filter(
            or_(
                KeyLog.ticket_mbs.ilike(like),
                KeyLog.visitor_name.ilike(like),
                KeyLog.visitor_company.ilike(like),
                KeyLog.rack_location.ilike(like),
            )
        )

    # Filter status
    if status in ("borrowed", "returned"):
        q = q.filter_by(status=status)

    # Filter tanggal (pakai time_in)
    if date_from_str:
        try:
            df = datetime.strptime(date_from_str, "%Y-%m-%d").date()
            start_dt = datetime.combine(df, datetime.min.time())
            q = q.filter(KeyLog.time_in >= start_dt)
        except ValueError:
            pass

    if date_to_str:
        try:
            dt_ = datetime.strptime(date_to_str, "%Y-%m-%d").date()
            end_dt = datetime.combine(dt_, datetime.max.time())
            q = q.filter(KeyLog.time_in <= end_dt)
        except ValueError:
            pass

    # Sort
    if sort == "time_in_asc":
        q = q.order_by(KeyLog.time_in.asc())
    else:  # default
        q = q.order_by(KeyLog.time_in.desc())

    return q.all()


@key_bp.route("/")
@login_required
def index():
    logs = get_filtered_logs(request.args)

    borrowed_logs = (
        KeyLog.query
        .filter(
            KeyLog.status == "borrowed",
            KeyLog.rack_location.in_(RACK_NAMES),
        )
        .order_by(KeyLog.rack_location.asc())
        .all()
    )

    return render_template(
        "key/index.html",
        logs=logs,
        borrowed_logs=borrowed_logs,
    )


@key_bp.route("/data")
@login_required
def data():
    logs = get_filtered_logs(request.args)
    return render_template("key/_table_body.html", logs=logs)


@key_bp.route("/borrow-form")
@login_required
def borrow_form():
    now = datetime.now()

    borrowed_logs = (
        KeyLog.query
        .filter(
            KeyLog.status == "borrowed",
            KeyLog.rack_location.in_(RACK_NAMES),
        )
        .all()
    )
    borrowed_set = {log.rack_location for log in borrowed_logs}

    racks = [
        {"name": name, "is_borrowed": name in borrowed_set}
        for name in RACK_NAMES
    ]

    return render_template(
        "key/_borrow_modal.html",
        now=now,
        racks=racks,
        staff_names=STAFF_NAMES,
    )


@key_bp.route("/borrow", methods=["POST"])
@login_required
def borrow():
    ticket_mbs = request.form.get("ticket_mbs")
    visitor_name = request.form.get("visitor_name")
    visitor_company = request.form.get("visitor_company") or ""
    rack_location = request.form.get("rack_location")
    time_in_str = request.form.get("time_in")
    staff_in_name = request.form.get("staff_in_name")   # ← NEW

    sig_in_visitor = request.form.get("sig_in_visitor")
    sig_in_staff = request.form.get("sig_in_staff")

    if not ticket_mbs or not visitor_name or not rack_location or not time_in_str or not staff_in_name:
        logs = get_filtered_logs(request.args)
        return render_template("key/_table_body.html", logs=logs), 400

    if not sig_in_visitor or not sig_in_staff:
        logs = get_filtered_logs(request.args)
        return render_template("key/_table_body.html", logs=logs), 400

    # validasi basic
    if not ticket_mbs or not visitor_name or not rack_location or not time_in_str:
        logs = KeyLog.query.order_by(KeyLog.created_at.desc()).all()
        return render_template("key/_table_body.html", logs=logs), 400

    # validasi TTD tambahan (backup server-side)
    if not sig_in_visitor or not sig_in_staff:
        logs = KeyLog.query.order_by(KeyLog.created_at.desc()).all()
        return render_template("key/_table_body.html", logs=logs), 400

    # CEK: rak ini lagi dipinjam gak?
    existing = KeyLog.query.filter_by(
        rack_location=rack_location,
        status="borrowed",
    ).first()
    if existing:
        # untuk sekarang kita cuma balikin tabel lagi dengan status 400
        # (nanti bisa ditambah notifikasi HTMX kalau mau)
        logs = get_filtered_logs(request.args)
        return render_template("key/_table_body.html", logs=logs), 400

    ############ Fingerprint verification ############
    fp = get_fingerprint_service()

    # 1️⃣ VERIFY / ENROLL PEMINJAM
    result_borrower = fp.verify(visitor_name)

    if not result_borrower.get("matched"):
        enroll_result = fp.enroll(visitor_name)
        borrower_fp_id = enroll_result["fingerprint_id"]
    else:
        borrower_fp_id = result_borrower["fingerprint_id"]



    # 2️⃣ VERIFY PETUGAS (yang menyerahkan kunci)
    result_staff = fp.verify(staff_in_name)
    if not result_staff.get("matched"):
        
        logs = get_filtered_logs(request.args)
        return render_template(
            "key/_borrow_modal.html",
            now=datetime.now(),
            logs=logs,
            staff_names=STAFF_NAMES,
            error_message="Fingerprint petugas belum terdaftar. Silakan enroll terlebih dahulu."
        ), 403
    staff_in_fp_id = result_staff["fingerprint_id"]
    

    ############ END OF FINGERPRINT VERIFICATION ############


    try:
        time_in = datetime.strptime(time_in_str, "%Y-%m-%dT%H:%M")
    except ValueError:
        time_in = datetime.now()

    log = KeyLog(
        ticket_mbs=ticket_mbs,
        visitor_name=visitor_name,
        visitor_company=visitor_company,
        rack_location=rack_location,
        time_in=time_in,
        status="borrowed",
        signature_in_visitor=sig_in_visitor,
        signature_in_staff=sig_in_staff,
        staff_in_name=staff_in_name,
        borrower_fp_id=borrower_fp_id,
        staff_in_fp_id=staff_in_fp_id,
        created_by_id=current_user.id if current_user.is_authenticated else None,
)


    db.session.add(log)
    db.session.commit()

    if request.headers.get("HX-Request"):
        logs = KeyLog.query.order_by(KeyLog.created_at.desc()).all()
        return render_template("key/_table_body.html", logs=logs)

    return redirect(url_for("key.index"))


@key_bp.route("/return-form/<int:log_id>")
@login_required
def return_form(log_id):
    log = KeyLog.query.get_or_404(log_id)
    if log.status != "borrowed":
        abort(400)

    now = datetime.now()
    return render_template(
        "key/_return_modal.html",
        log=log,
        now=now,
        staff_names=STAFF_NAMES,
    )


@key_bp.route("/return/<int:log_id>", methods=["POST"])
@login_required
def return_key(log_id):
    log = KeyLog.query.get_or_404(log_id)
    if log.status != "borrowed":
        abort(400)

    time_out_str = request.form.get("time_out")
    sig_out_visitor = request.form.get("sig_out_visitor")
    sig_out_staff = request.form.get("sig_out_staff")
    staff_out_name = request.form.get("staff_out_name")

    if not time_out_str or not staff_out_name:
        logs = get_filtered_logs(request.args)
        return render_template("key/_table_body.html", logs=logs), 400

    if not sig_out_visitor or not sig_out_staff:
        logs = get_filtered_logs(request.args)
        return render_template("key/_table_body.html", logs=logs), 400

    if not time_out_str:
        logs = KeyLog.query.order_by(KeyLog.created_at.desc()).all()
        return render_template("key/_table_body.html", logs=logs), 400

    # validasi TTD
    if not sig_out_visitor or not sig_out_staff:
        logs = KeyLog.query.order_by(KeyLog.created_at.desc()).all()
        return render_template("key/_table_body.html", logs=logs), 400

    try:
        time_out = datetime.strptime(time_out_str, "%Y-%m-%dT%H:%M")
    except ValueError:
        time_out = datetime.now()

    # ===== Fingerprint verification (RETURN) =====
    fp = get_fingerprint_service()

    # 1️⃣ VERIFY PEMINJAM (yang mengembalikan)
    result_borrower = fp.verify(log.visitor_name)
    
    if not result_borrower.get("matched"):
        logs = get_filtered_logs(request.args)
        return render_template(
            "key/_return_modal.html",
            log=log,
            now=datetime.now(),
            staff_names=STAFF_NAMES,
            error_message="Fingerprint peminjam tidak valid. Silakan scan ulang."
        ), 403
    borrower_fp_id = result_borrower["fingerprint_id"]
    



    # 2️⃣ VERIFY PETUGAS (yang menerima, shift bebas)
    result_staff = fp.verify(staff_out_name)
    if not result_staff.get("matched"):
        logs = get_filtered_logs(request.args)
        return render_template(
            "key/_return_modal.html",
            log=log,
            now=datetime.now(),
            staff_names=STAFF_NAMES,
            error_message="Fingerprint petugas belum terdaftar. Silakan enroll terlebih dahulu."
        ), 403

    staff_out_fp_id = result_staff["fingerprint_id"]
    log.staff_out_fp_id = staff_out_fp_id

    # ===== END OF Fingerprint verification =====



    log.time_out = time_out
    log.status = "returned"
    log.signature_out_visitor = sig_out_visitor
    log.signature_out_staff = sig_out_staff
    log.staff_out_name = staff_out_name  # ← NEW
    log.updated_by_id = current_user.id if current_user.is_authenticated else None
    log.borrower_fp_id = borrower_fp_id

    db.session.commit()

    if request.headers.get("HX-Request"):
        logs = KeyLog.query.order_by(KeyLog.created_at.desc()).all()
        return render_template("key/_table_body.html", logs=logs)

    return redirect(url_for("key.index"))

@key_bp.route("/detail/<int:log_id>")
@login_required
def detail(log_id):
    log = KeyLog.query.get_or_404(log_id)
    return render_template("key/_detail_modal.html", log=log)

@key_bp.route("/export/pdf")
@login_required
def export_pdf():
    logs = get_filtered_logs(request.args)

    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    def draw_header():
        c.setFont("Helvetica-Bold", 14)
        c.drawString(20 * mm, height - 20 * mm, "Laporan Peminjaman Kunci Rak")
        c.setFont("Helvetica", 9)
        c.drawString(20 * mm, height - 26 * mm, f"Dibuat: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        c.line(20 * mm, height - 28 * mm, width - 20 * mm, height - 28 * mm)

    def draw_signature_box(label, sig_data, x, y_box, box_width, box_height):
        c.setFont("Helvetica", 8)
        c.drawString(x, y_box + box_height + 2 * mm, label)
        c.rect(x, y_box, box_width, box_height)
        if not sig_data:
            c.setFont("Helvetica-Oblique", 7)
            c.drawString(x + 2 * mm, y_box + box_height / 2, "Tidak ada TTD.")
            return
        try:
            if sig_data.startswith("data:image"):
                sig_data_base64 = sig_data.split(",", 1)[1]
            else:
                sig_data_base64 = sig_data
            img_bytes = base64.b64decode(sig_data_base64)
            img_reader = ImageReader(io.BytesIO(img_bytes))
            c.drawImage(img_reader, x + 2 * mm, y_box + 2 * mm,
                        width=box_width - 4 * mm, height=box_height - 4 * mm,
                        preserveAspectRatio=True, anchor="c")
        except Exception:
            c.setFont("Helvetica-Oblique", 7)
            c.drawString(x + 2 * mm, y_box + box_height / 2, "TTD tidak dapat ditampilkan.")

    draw_header()
    y = height - 40 * mm

    if not logs:
        c.setFont("Helvetica-Oblique", 10)
        c.drawString(20 * mm, y, "Tidak ada transaksi pada filter yang dipilih.")
    else:
        bottom_margin = 30 * mm  # jarak aman dari bawah

        for idx, log in enumerate(logs, start=1):
            # kalau posisi Y udah terlalu rendah, pindah halaman baru
            if y < bottom_margin:
                c.showPage()
                draw_header()
                y = height - 40 * mm

            c.setFont("Helvetica-Bold", 10)
            c.drawString(20 * mm, y, f"{idx}. Ticket: {log.ticket_mbs}")
            y -= 5 * mm

            c.setFont("Helvetica", 9)
            visitor_line = f"Pengunjung: {log.visitor_name}"
            if log.visitor_company:
                visitor_line += f" ({log.visitor_company})"
            c.drawString(20 * mm, y, visitor_line)
            y -= 4 * mm

            c.drawString(20 * mm, y, f"Lokasi Rak: {log.rack_location}")
            y -= 4 * mm
            c.drawString(20 * mm, y, f"Status: {log.status}")
            y -= 4 * mm
            c.drawString(20 * mm, y, f"Petugas (PINJAM): {log.staff_in_name or '-'}")
            y -= 4 * mm
            c.drawString(20 * mm, y, f"Petugas (KEMBALI): {log.staff_out_name or '-'}")
            y -= 4 * mm
            c.drawString(20 * mm, y, f"Jam Masuk: {log.time_in.strftime('%Y-%m-%d %H:%M') if log.time_in else '-'}")
            y -= 4 * mm
            c.drawString(20 * mm, y, f"Jam Keluar: {log.time_out.strftime('%Y-%m-%d %H:%M') if log.time_out else '-'}")
            y -= 6 * mm

            # ukuran kotak ttd
            box_width = (width - 40 * mm) / 2
            box_height = 25 * mm

            # posisi pertama (pinjam)
            y_box1 = y - box_height
            if y_box1 < bottom_margin:
                # kalau kotak ga muat, new page dulu
                c.showPage()
                draw_header()
                y = height - 40 * mm
                y_box1 = y - box_height

            draw_signature_box("TTD Pengunjung (PINJAM)", log.signature_in_visitor, 20 * mm, y_box1, box_width, box_height)
            draw_signature_box("TTD Petugas (PINJAM)", log.signature_in_staff, 20 * mm + box_width, y_box1, box_width, box_height)
            y = y_box1 - 7 * mm

            # posisi kedua (kembali)
            y_box2 = y - box_height
            if y_box2 < bottom_margin:
                c.showPage()
                draw_header()
                y = height - 40 * mm
                y_box2 = y - box_height

            draw_signature_box("TTD Pengunjung (KEMBALI)", log.signature_out_visitor, 20 * mm, y_box2, box_width, box_height)
            draw_signature_box("TTD Petugas (KEMBALI)", log.signature_out_staff, 20 * mm + box_width, y_box2, box_width, box_height)
            y = y_box2 - 10 * mm  # jarak antar transaksi


    c.save()
    pdf = buffer.getvalue()
    buffer.close()

    filename = f"RACKess_Peminjaman_Kunci_Filter_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    response = make_response(pdf)
    response.headers["Content-Type"] = "application/pdf"
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response


@key_bp.route("/export/excel")
@login_required
def export_excel():
    # pakai helper yang sama dengan tabel (ikut filter + search + tanggal)
    logs = get_filtered_logs(request.args)

    wb = Workbook()
    ws = wb.active
    ws.title = "Peminjaman Kunci"

    # header
    headers = [
        "ID",
        "Ticket MBS",
        "Pengunjung",
        "Perusahaan / Divisi",
        "Lokasi Rak",
        "Status",
        "Jam Masuk",
        "Jam Keluar",
        "Petugas PINJAM",
        "Petugas KEMBALI",
    ]
    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        
    for log in logs:
        ws.append([
            log.id,
            log.ticket_mbs,
            log.visitor_name,
            log.visitor_company or "",
            log.rack_location,
            log.status,
            log.time_in.strftime("%Y-%m-%d %H:%M") if log.time_in else "",
            log.time_out.strftime("%Y-%m-%d %H:%M") if log.time_out else "",
            log.staff_in_name or "",
            log.staff_out_name or "",
        ])


    # === AUTO-FIT LEBAR KOLOM ===
    from openpyxl.utils import get_column_letter
    for column_cells in ws.columns:
        max_length = 0
        column = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            try:
                # hitung panjang teks
                cell_length = len(str(cell.value)) if cell.value else 0
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width


    # bikin file di memory
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    # nama file pakai tanggal filter kalau ada
    date_from = (request.args.get("date_from") or "").strip()
    date_to = (request.args.get("date_to") or "").strip()
    if date_from or date_to:
        suffix = f"{date_from or 'all'}_to_{date_to or 'all'}"
    else:
        suffix = "all"

    filename = f"RACKess_Peminjaman_Kunci_{suffix}.xlsx"

    response = make_response(file_stream.getvalue())
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response
